from django.shortcuts import render, redirect, get_object_or_404
from .models import Employees, Sites, SiteManagers, Timesheets
from django.db.models import Q
from control_panel.utils import platform_admin_required, tenant_admin_required, site_manager_required
from p.p.routers import set_current_tenant
from control_panel.models import PublicHolidays, Company, Tenant
from .forms import EmployeeForm, SiteForm, SiteManagersForm, TimeSheetsForm, UploadFileForm
# from .routers import set_current_user
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import FileResponse, HttpResponse
from django.conf import settings
from django.utils.dateparse import parse_datetime
from django.template.loader import render_to_string
from collections import Counter, defaultdict
import os
from uuid import uuid4
import openpyxl
from openpyxl import Workbook
# import datetime
from datetime import datetime, timedelta, date as dt_date, time as dt_time
from xhtml2pdf import pisa
from django.utils import timezone
import math
import pandas as pd

# Create your views here.


def superuser_required(view_func):
    decorated_view_func = login_required(user_passes_test(lambda u: u.is_superuser)(view_func))
    return decorated_view_func

def is_superuser(user):
    return user.is_superuser



#
def logout_view(request):
    logout(request)
    return redirect('index')

def media_files(request, path):
    # Construct the full path to the media file
    full_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(full_path):
        # Return the media file response
        return FileResponse(open(full_path, 'rb'))
    else:
        # Return a 404 error if the file doesn't exist
        from django.http import Http404
        raise Http404("File does not exist")


def index(request):

    return render(request,"control_panel/index.html",)


def get_user_site_id(user):
    logged_in_user_site = get_object_or_404(SiteManagers, user_id=user.id).site
    return logged_in_user_site



@tenant_admin_required
def sites(request):
    sites = Sites.objects.filter(company_id=request.user.company_id)

    return render(request, "timesheets/sites.html", {
        "sites": sites
    })





def profile(request,user):
    user = request.user

    if is_superuser(user):
        img = 'https://upload.wikimedia.org/wikipedia/commons/5/55/User-admin-gear.svg'
        return img
    else:
        return 'https://storage.needpix.com/rsynced_images/user-1633250_1280.png'

@login_required
def timesheets(request):
    user = request.user
    if is_superuser(user) or user.role in ["tenant_admin", "platform_admin"]:
        sites = list(Sites.objects.filter(company_id=user.company_id))  # mutable list

        # Find unsigned timesheets with no valid site
        existing_site_ids = Sites.objects.filter(company_id=user.company_id).values_list('site_id', flat=True)
        unknown_timesheets = Timesheets.objects.filter(
            company_id=user.company_id,
            signed_off=None
        ).exclude(
            site_id__in=existing_site_ids
        )

        if unknown_timesheets.exists():
            unknown_site = Sites(site_id=None, site_name="Unknown Site", company_id=user.company_id)
            unknown_site.is_unknown = True
            sites.append(unknown_site)

        today = datetime.today().date()
        return render(request, "timesheets/timesheets.html", {
            "sites": Sites.objects.filter(company_id=user.company_id),
            "today": today,
            "sites_count": get_site_to_signed_off_timesheets(request),
            "unknown_count": unknown_timesheets.count()
        })

    else:
        user_site = get_user_site_id(user)
        sites = Sites.objects.filter(site_id=user_site.site_id, company_id=user.company_id)
        today = datetime.today().date()
        is_site_manager = user.groups.filter(name='SiteManager').exists()

        return render(request, "timesheets/timesheets.html", {
            "sites": sites,
            "today": today,
            "is_site_manager": is_site_manager,
            "sites_count": get_site_to_signed_off_site_manager(user_site.site_id, request)
        })


def get_site_to_signed_off_site_manager(site_id, request):
    # Query to get the site IDs where signed_off is None
    non_signed_off = Timesheets.objects.filter(signed_off=None, site_id=site_id, company_id=request.user.company_id).values('site_id')
    count = Timesheets.objects.filter(signed_off=None, clock_out__isnull=False, site_id=site_id, company_id=request.user.company_id).count()


    # Get all sites (including site_name and site_id)
    curret_site = Sites.objects.filter(site_id=site_id, company_id=request.user.company_id).values('site_id', 'site_name', 'site_address')



    # Initialize a list to store the final result
    site_info_list = []

    # Populate the list with site_name, site_id, and site_count
    for site in curret_site:
        site_id = site['site_id']
        site_name = site['site_name']
        site_address = site['site_address']
        site_count = count
        site_info_list.append({
            'site_name': site_name,
            'site_id': site_id,
            'site_count': site_count,
            'site_address': site_address,
            'company': request.user.company_id
        })
    return site_info_list

def get_site_to_signed_off_timesheets(request):
    # Query to get the site IDs where signed_off is None
    non_signed_off = Timesheets.objects.filter(signed_off=None, company_id=request.user.company_id).values('site_id', 'clock_out')
    site_ids = [site['site_id'] for site in non_signed_off if site['clock_out'] != None]

    # Get all sites (including site_name and site_id)
    all_sites = Sites.objects.filter(company_id=request.user.company_id).values('site_id', 'site_name', 'site_address')

    # Count the occurrences of each site_id
    site_id_counts = Counter(site_ids)

    # Initialize a list to store the final result
    site_info_list = []

    # Populate the list with site_name, site_id, and site_count
    for site in all_sites:

        site_id = site['site_id']
        site_name = site['site_name']
        site_address = site['site_address']
        site_count = site_id_counts.get(site_id, 0)
        site_info_list.append({
            'site_name': site_name,
            'site_id': site_id,
            'site_count': site_count,
            'site_address': site_address
        })

    return site_info_list


@tenant_admin_required
def add_site(request):
    form = SiteForm(initial={'site_id':  str(uuid4()).split('-')[2],
                             'created_date': datetime.now()})
    if request.method == "POST":
        form = SiteForm(request.POST, user=request.user)
        if form.is_valid():
            site_id = form.cleaned_data['site_id']

            if Sites.objects.filter(site_id=site_id).exists():
                messages.success(request, f'site already Added')
                return redirect('sites')
            else:
                form.save()
                messages.success(request, 'New staff added')
                return redirect("sites")

    # context = {'form': form}
    return render(request, "timesheets/add_site.html", {
        'form': form
    })
@login_required
def get_employee_list(employees, site):
    employees_list = []
    site_managers_ids = SiteManagers.objects.filter(site_id=site).values_list('employee_id', flat=True)

    for employee in employees:
        if employee.employee_id not in site_managers_ids:
            employees_list.append(employee)

    return employees_list

@tenant_admin_required
def add_site_managers(request, site_id):
    site_instance = get_object_or_404(Sites, site_id=site_id)
    if request.method == "POST":
        form = SiteManagersForm(request.POST, user=request.user, site_instance=site_instance,initial={'manager_site_id':  str(uuid4()).split('-')[2],})
        if form.is_valid():
            form.save()
            return redirect('view_site', site_id=site_id)
    else:
        form = SiteManagersForm(user=request.user, initial={'site': site_instance, 'assigned': datetime.now()}, site_instance=site_instance)

    return render(request, "timesheets/add_site_managers.html", {
        'form': form,
        'site': site_instance
    })


from django.shortcuts import get_object_or_404

@tenant_admin_required
def view_site(request, site_id):
    # Get the site object (more robust than filtering by raw id)
    site = get_object_or_404(Sites, site_id=site_id)

    # Build list of dicts: {employee_id, id (site_manager_id)}
    site_managers = SiteManagers.objects.filter(site=site)
    site_managers_info = [
        {"employee_id": sm.employee_id, "id": sm.id}
        for sm in site_managers
    ]

    employees = Employees.objects.all()

    employees_info_list = []
    for employee in employees:
        # Determine the key used in SiteManagers: pk or business employee_id
        possible_keys = [employee.id, getattr(employee, "employee_id", None)]

        # Find a match in the list-of-dicts (works even if employee_id is str or int)
        match = next(
            (row for row in site_managers_info if row["employee_id"] in possible_keys),
            None,
        )

        if match:
            employees_info_list.append({
                "employee_id": getattr(employee, "employee_id", employee.id),
                "first_name": employee.first_name,
                "last_name": employee.last_name,
                "position": getattr(employee, "position", None),
                "site_manager_id": match["id"],  # <-- the SiteManagers.id you wanted
            })

    site_employees_count = Employees.objects.filter(site=site).count()

    return render(request, "timesheets/view_site.html", {
        "site": site,
        "site_managers": site_managers,
        "site_employees": site_employees_count,
        "employee_info": employees_info_list,
    })


@tenant_admin_required
def view_site_managers(request, site_id):
    site = Sites.objects.get(site_id=site_id)
    site_managers = SiteManagers.objects.filter(site_id=site_id)
    return render(request, "timesheets/view_site_managers.html", {
        'site': site,
        'site_managers': site_managers
    })

@tenant_admin_required
def view_site_manager(request, site_id, id):
    site = Sites.objects.get(site_id=site_id)
    site_manager = SiteManagers.objects.get(id=id)
    employee_id = site_manager.employee_id
    employee = Employees.objects.get(id=employee_id)
    return render(request, "timesheets/view_site_manager.html", {
        'site': site,
        'site_manager': site_manager,
        'employee': employee
    })

@tenant_admin_required
def delete_site_manager(request, site_id, id):
    site_manager = SiteManagers.objects.get(id=id)

    if site_manager and site_manager.user:  # Check if the SiteManager has an associated user
        user = site_manager.user  # Get the associated User
        site_manager.delete()  # Delete the SiteManager
        user.delete()  # Explicitly delete the associated User

    messages.success(request, 'Site Manager and associated User removed.')
    return redirect('view_site', site_id=site_id)


@tenant_admin_required
def edit_site(request, site_id):
    site = Sites.objects.get(site_id=site_id)
    form = SiteForm(instance=site)
    if request.method == "POST":
        form = SiteForm(request.POST, instance=site)
        if form.is_valid():
            form.save()
            messages.success(request, f"{site} edited")
            return redirect('view_site',  site_id=site_id)

    context = {'form': form,
               'site': site
}
    return render(request, 'timesheets/edit_site.html', context)

@tenant_admin_required
def delete_site(request, site_id):
    site = Sites.objects.get( site_id=site_id)
    site.delete()
    messages.success(request, f'Site Removed')
    return redirect('sites')


@tenant_admin_required
def employees(request):
    # Retrieve the selected site and search query from the GET parameters
    selected_site = request.GET.get("site")
    search_query = request.GET.get("search")

    # ✅ Filter employees by logged-in user's company
    employees = Employees.objects.filter(company_id=request.user.company_id)

    # Optional: further filter by site
    if selected_site:
        employees = employees.filter(site_id=selected_site)

    # Optional: further filter by search query
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
        )

    employees_info_list = []
    for employee in employees:
        site_name = employee.site if employee.site else "Unknown"
        employees_info_list.append({
            "id": employee.id,
            "employee_id": employee.employee_id,
            "first_name": employee.first_name,
            "last_name": employee.last_name,
            "connected": employee.connected,
            "site": site_name,
        })

    # ✅ Restrict sites dropdown to the same company too
    sites = Sites.objects.filter(company_id=request.user.company_id)

    return render(request, "timesheets/employees.html", {
        "employees": employees_info_list,
        "sites": sites,
        "selected_site": selected_site,
        "search_query": search_query,
    })

@tenant_admin_required
def add_employee(request):
    user = request.user
    if request.method == "POST":
        form = EmployeeForm(request.POST, request.FILES, user=request.user, initial={'company': user.company})
        if form.is_valid():
            employee_id = form.cleaned_data['employee_id']
            id_number = form.cleaned_data['id_number']
            if Employees.objects.filter(employee_id=employee_id, company_id=request.user.company_id, id_number=id_number ).exists():
                messages.error(request, 'User already exists')
                return redirect('index')
            else:
                form.save()
                messages.success(request, 'New employee added')
                return redirect("employees")
    else:
        form = EmployeeForm(user=request.user)  # Correctly instantiate the form when GET request

    return render(request, "timesheets/add_employee.html", {
        'form': form,

    })

@tenant_admin_required
def view_employee(request, id):
    employee = Employees.objects.get(id=id)
    return render(request, "timesheets/view_employee.html", {
        'employee': employee
    })
@tenant_admin_required
def edit_employee(request, id):
    employee = Employees.objects.get(id=id)
    if request.method == "POST":

        form = EmployeeForm(request.POST, request.FILES, instance=employee, user=request.user)  # Pass request.FILES here
        if form.is_valid():
            form.save()
            messages.success(request, f"{employee.first_name} {employee.last_name} has been edited")
            return redirect('employees', )
    else:
        form = EmployeeForm(instance=employee, user=request.user)  # Correctly instantiate the form when GET request

    return render(request, "timesheets/edit_employee.html", {
        'form': form,
        'employee': employee
    })

@tenant_admin_required
def delete_employee(request, id, employee_id):
    try:
        # Get the employee object using the primary key (id)
        employee = Employees.objects.get(id=id)

        # Get all timesheets related to this employee using the employee_id
        employee_timesheets = Timesheets.objects.filter(employee_id=employee_id)

        # Delete all related timesheets
        employee_timesheets.delete()

        # Now delete the employee
        employee.delete()

        # Add a success message to the session
        messages.success(request, f'Employee and related timesheets removed successfully')

    except Employees.DoesNotExist:
        # If the employee doesn't exist, show an error message
        messages.error(request, f'Employee does not exist')

    # Redirect back to the employees page
    return redirect('employees')


@login_required
def view_site_employees(request, site_id):
    employees = get_employee_list(site_id, request)
    site = Sites.objects.get(site_id=site_id)

    return render(request, "timesheets/view_site_employees.html", {
        "employees": employees,
        'employees_list': get_employee_list(site_id, request),

        "site": site,
        "today": datetime.today().date(),  # Pass today's date to the template
    })

def get_employee_list(site_id, request):

    non_signed_off = Timesheets.objects.filter(signed_off=None, site_id=site_id, company_id=request.user.company_id).values('employee_id', 'clock_out')
    # count = Timesheets.objects.filter(signed_off=None, site_id=site_id)
    employee_ids = [employee['employee_id'] for employee in non_signed_off if employee['clock_out'] != None]
    site_employees= Employees.objects.filter(company_id=request.user.company_id ).values('employee_id','first_name', 'last_name', 'connected','position', 'id')

    # Count the occurrences of each employee_id
    employee_id_counts = Counter(employee_ids)

    employees_info_list = []


    for employee in site_employees:
        id = employee['id']
        employee_id = employee['employee_id']
        count = employee_id_counts.get(employee_id, 0)
        if count == 0:
            continue

        employee_first_name = employee['first_name']
        employee_last_name = employee['last_name']
        employee_connected = employee['connected']
        employee_position = employee['position']

        employees_info_list.append({
            'id': id,
            'employee_id': employee_id,
            'first_name': employee_first_name,
            'last_name': employee_last_name,
            'connected': f'{employee_connected}',
            'position': employee_position,
            'count': count

        })

    return employees_info_list

@tenant_admin_required
def view_unknown_site_employees(request):
    # Get all unsigned timesheets with no site
    non_signed_off = Timesheets.objects.filter(signed_off=None, site__isnull=True, company_id=request.user.company_id).values('employee_id')

    # Extract unique employee IDs
    employee_ids = {t['employee_id'] for t in non_signed_off}

    # Get employee details for those IDs
    employees = Employees.objects.filter(employee_id__in=employee_ids, company_id=request.user.company_id)

    # Reuse get_employee_list logic but adapt for "Unknown Site"
    employees_list = get_employee_list_for_unknown_site(request)

    return render(request, "timesheets/unknown_site_timesheets.html", {
        "employees": employees_list,
        "employees_list": employees_list,
        "site": {"site_name": "Unknown Site"},  # Fake site object/dict for template display
        "today": datetime.today().date(),
    })


def get_employee_list_for_unknown_site(request):
    # Get unsigned timesheets without a site
    non_signed_off = Timesheets.objects.filter(signed_off=None, site__isnull=True, company_id=request.user.company_id).values('employee_id')

    employee_ids = [employee['employee_id'] for employee in non_signed_off]
    site_employees = Employees.objects.filter(company_id=request.user.company_id).values('employee_id', 'first_name', 'last_name', 'connected', 'position', 'id')

    # Count occurrences of each employee_id
    employee_id_counts = Counter(employee_ids)

    employees_info_list = []

    for employee in site_employees:
        id = employee['id']
        employee_id = employee['employee_id']
        count = employee_id_counts.get(employee_id, 0)
        if count == 0:
            continue

        employees_info_list.append({
            'id': id,
            'employee_id': employee_id,
            'first_name': employee['first_name'],
            'last_name': employee['last_name'],
            'connected': f"{employee['connected']}",
            'position': employee['position'],
            'count': count
        })

    return employees_info_list


@tenant_admin_required
def payroll_data(request):
    # Get the start and end dates from the request, with defaults if not provided
    start_date_str = request.GET.get('start_date', '').strip()  # Use .strip() to remove any leading/trailing whitespace
    end_date_str = request.GET.get('end_date', '').strip()

    # Set default values if the date strings are empty
    if not start_date_str:
        start_date_str = '2025-01-01'
    if not end_date_str:
        end_date_str = '2025-12-31'

    # Convert the date strings to datetime objects
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except ValueError as e:
        # Handle invalid date format
        return render(request, "timesheets/error.html", {"message": f"Invalid date format: {e}"})

    # Filter the timesheets that have been signed off and are within the date range
    timesheets = Timesheets.objects.filter(date__range=[start_date, end_date], signed_off='Yes', company_id=request.user.company_id)

    # Create dictionaries to store the total hours worked for each employee
    employee_hours = defaultdict(float)
    employee_normal_hours = defaultdict(float)
    employee_overtime_normal_saturdays = defaultdict(float)

    employee_sunday_hours = defaultdict(float)
    employee_public_holiday_hours = defaultdict(float)

    for timesheet in timesheets:
        employee_id = timesheet.employee_id
        employee_hours[employee_id] += timesheet.hours_worked
        employee_normal_hours[employee_id] += timesheet.normal_hours
        employee_overtime_normal_saturdays[employee_id] += timesheet.overtime_normal_saturdays

        # Separate Sunday vs Public Holiday hours
        if timesheet.date.weekday() == 6:  # Sunday
            employee_sunday_hours[employee_id] += timesheet.overtime_holiday_sundays
        elif is_public_holiday(timesheet.date):
            employee_public_holiday_hours[employee_id] += timesheet.overtime_holiday_sundays

    employees = Employees.objects.filter(company_id=request.user.company_id).values(
        'employee_id', 'first_name', 'last_name', 'position', 'site', 'wage'
    )

    employee_payroll_data = []
    for employee in employees:
        employee_id = employee['employee_id']
        total_hours_worked = round(employee_hours.get(employee_id, 0), 2)
        normal_hours = employee_normal_hours.get(employee_id, 0)
        overtime_one_hours = employee_overtime_normal_saturdays.get(employee_id, 0)
        sunday_hours = employee_sunday_hours.get(employee_id, 0)
        public_holiday_hours = employee_public_holiday_hours.get(employee_id, 0)

        if total_hours_worked > 0:
            employee_info = {
                'employee_id': employee_id,
                'first_name': employee['first_name'],
                'last_name': employee['last_name'],
                'position': employee['position'],
                'total_hours_worked': total_hours_worked,
                'total_normal_hours': round(normal_hours, 2),
                'total_overtime_normal_saturdays': overtime_one_hours,
                'total_overtime_holiday_sundays': sunday_hours + public_holiday_hours,

                'site': Sites.objects.get(site_id=employee['site']).site_name,
                'owed': get_owed_amount(
                    request,
                    employee['wage'],
                    normal_hours,
                    overtime_one_hours,
                    sunday_hours,
                    public_holiday_hours
                ),
                'period': f'{format_date(start_date)} - {format_date(end_date)}'
            }
            employee_payroll_data.append(employee_info)

    return render(request, "timesheets/pay_roll_data_view.html", {
        'employees': employee_payroll_data,  # Use the filtered list
        'timesheets': timesheets,
        'start_date': start_date_str,
        'end_date': end_date_str,


    })



def render_to_pdf(template_src, context_dict={}, filename=None):
    template = render_to_string(template_src, context_dict)
    response = HttpResponse(content_type='application/pdf')

    # Set the default filename if none is provided
    if not filename:
        filename = 'payroll_data.pdf'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Create PDF
    pisa_status = pisa.CreatePDF(template, dest=response)
    if pisa_status.err:
        return HttpResponse('Error creating PDF', status=400)
    return response

@tenant_admin_required
def payroll_data_pdf(request, start_date, end_date):
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError as e:
        return render(request, "timesheets/error.html", {"message": f"Invalid date format: {e}"})

    # Filter timesheets by date range and signed-off status
    timesheets = Timesheets.objects.filter(
        date__range=[start_date, end_date],
        signed_off='Yes',
        company_id=request.user.company_id
    )

    # Dictionaries for tracking
    employee_hours = defaultdict(float)
    employee_normal_hours = defaultdict(float)
    employee_overtime_normal_saturdays = defaultdict(float)
    employee_sunday_hours = defaultdict(float)
    employee_public_holiday_hours = defaultdict(float)

    # Function to check if a date is a public holiday (see below for the helper)

    for timesheet in timesheets:
        employee_id = timesheet.employee_id
        employee_hours[employee_id] += timesheet.hours_worked
        employee_normal_hours[employee_id] += timesheet.normal_hours
        employee_overtime_normal_saturdays[employee_id] += timesheet.overtime_normal_saturdays

        # Separate Sunday vs Public Holiday hours
        if timesheet.date.weekday() == 6:  # Sunday
            employee_sunday_hours[employee_id] += timesheet.overtime_holiday_sundays
        elif is_public_holiday(timesheet.date):
            employee_public_holiday_hours[employee_id] += timesheet.overtime_holiday_sundays

    employees = Employees.objects.filter(company_id=request.user.company_id).values(
        'employee_id', 'first_name', 'last_name', 'position', 'site', 'wage'
    )

    employee_payroll_data = []
    for employee in employees:
        employee_id = employee['employee_id']
        total_hours_worked = round(employee_hours.get(employee_id, 0), 2)
        normal_hours = employee_normal_hours.get(employee_id, 0)
        overtime_one_hours = employee_overtime_normal_saturdays.get(employee_id, 0)
        sunday_hours = employee_sunday_hours.get(employee_id, 0)
        public_holiday_hours = employee_public_holiday_hours.get(employee_id, 0)

        if total_hours_worked > 0:
            employee_info = {
                'employee_id': employee_id,
                'first_name': employee['first_name'],
                'last_name': employee['last_name'],
                'position': employee['position'],
                'total_hours_worked': total_hours_worked,
                'total_normal_hours': round(normal_hours, 2),
                'total_overtime_normal_saturdays': overtime_one_hours,
                'total_overtime_holiday_sundays': sunday_hours + public_holiday_hours,

                'site': Sites.objects.get(site_id=employee['site']).site_name,
                'owed': get_owed_amount(
                    request,
                    employee['wage'],
                    normal_hours,
                    overtime_one_hours,
                    sunday_hours,
                    public_holiday_hours
                ),
                'period': f'{format_date(start_date)} - {format_date(end_date)}'
            }
            employee_payroll_data.append(employee_info)

    filename = f'{format_date(start_date)} - {format_date(end_date)}_Payroll_Data.pdf'
    pdf = render_to_pdf('timesheets/pay_roll_data_pdf.html', {'employees': employee_payroll_data}, filename=filename)
    return pdf



def format_date(date):
    return date.strftime('%d %b %y')





@login_required
def view_employee_timesheets(request, id, site_id):
    current_year_month = get_year_month(datetime.today().date())
    timesheets = Timesheets.objects.filter(company_id=request.user.company_id, site_id=site_id, clock_out__isnull=False)
    employee = Employees.objects.get(id=id)
    site_name = site_id
    site = site_id
    timesheets_signed_off = Timesheets.objects.filter(employee_id=employee.employee_id, signed_off ='Yes', company_id=request.user.company_id)
    timesheets_to_signed_off = Timesheets.objects.filter(employee_id=employee.employee_id, signed_off=None, company_id=request.user.company_id, clock_out__isnull=False)
    year_month_list = []
    for year_month in timesheets_signed_off:
        date = get_year_month(year_month.date)
        year_month_list.append(date)

    timesheets_already_signed = [
        ts for ts in timesheets_signed_off if get_year_month(ts.date) == current_year_month
    ]
    signed_off_count = len(timesheets_already_signed)

    curren_month_timesheets = [
        ts for ts in timesheets if get_year_month(ts.date) == current_year_month
    ]
    total_hours_worked = sum([hour.hours_worked for hour in curren_month_timesheets  ])
    # Initialize a list to store the final result
    timesheets_info_list = []


    # Populate the list with employee details and total hours worked
    for timesheet in timesheets_to_signed_off:
        timesheets_info_list.append({
            'id': timesheet.id,
            'date': timesheet.date,
            'clock_in': timesheet.clock_in,
            'clock_out': timesheet.clock_out,
            'hours_worked': timesheet.hours_worked,
            'normal_hours': timesheet.normal_hours,
            'overtime_normal_saturdays': timesheet.overtime_normal_saturdays,

            'overtime_holiday_sundays': timesheet.overtime_holiday_sundays,
            'Day': is_it_saturday_sunday(timesheet.date),

        })


    return render(request, "timesheets/view_employee_timesheets.html", {
        "timesheets": timesheets_to_signed_off,
        'employee': employee,
        'signed_off' : timesheets_already_signed,
        'year_month': current_year_month,
        'year_month_list': year_month_list,
        'total_hours_worked' : round(total_hours_worked, 2),
        'timesheets_info': timesheets_info_list,
        'site': site,
        'signed_off_count': signed_off_count


        # Pass today's date to the template
    })


def view_employee_unknown_timesheets(request, id, employee_id ):
    current_year_month = get_year_month(datetime.today().date())
    timesheets = Timesheets.objects.filter( employee_id=employee_id, company_id=request.user.company_id, clock_out__isnull=False)
    employee = Employees.objects.get(id=id)
    timesheets_signed_off = Timesheets.objects.filter(employee_id=employee_id, signed_off ='Yes', company_id=request.user.company_id)
    timesheets_to_signed_off = Timesheets.objects.filter(employee_id=employee_id, signed_off=None, company_id=request.user.company_id)
    year_month_list = []
    for year_month in timesheets_signed_off:
        date = get_year_month(year_month.date)
        year_month_list.append(date)

    timesheets_already_signed = [
        ts for ts in timesheets_signed_off if get_year_month(ts.date) == current_year_month
    ]
    signed_off_count = len(timesheets_already_signed)

    curren_month_timesheets = [
        ts for ts in timesheets if get_year_month(ts.date) == current_year_month
    ]
    total_hours_worked = sum([hour.hours_worked for hour in curren_month_timesheets  ])
    # Initialize a list to store the final result
    timesheets_info_list = []


    # Populate the list with employee details and total hours worked
    for timesheet in timesheets_to_signed_off:
        timesheets_info_list.append({
            'id': timesheet.id,
            'date': timesheet.date,
            'clock_in': timesheet.clock_in,
            'clock_out': timesheet.clock_out,
            'hours_worked': timesheet.hours_worked,
            'normal_hours': timesheet.normal_hours,
            'overtime_normal_saturdays': timesheet.overtime_normal_saturdays,

            'overtime_holiday_sundays': timesheet.overtime_holiday_sundays,
            'Day': is_it_saturday_sunday(timesheet.date),

        })


    return render(request, "timesheets/view_employee_unknown_timesheets.html", {
        "timesheets": timesheets_to_signed_off,
        'employee': employee,
        'signed_off' : timesheets_already_signed,
        'year_month': current_year_month,
        'year_month_list': year_month_list,
        'total_hours_worked' : round(total_hours_worked, 2),
        'timesheets_info': timesheets_info_list,
        'signed_off_count': signed_off_count


        # Pass today's date to the template
    })



def get_month_number(month):
    if month == "Jan":
        return 1
    if month == "Feb":
        return 2
    if month == "Mar":
        return 3
    if month == "Apr":
        return 4
    if month == "May":
        return 5
    if month == "Jun":
        return 6
    if month == "Jul":
        return 7
    if month == "Aug":
        return 8
    if month == "Sept":
        return 9
    if month == "Oct":
        return 10
    if month == "Nov":
        return 11
    if month == "Dec":
        return 12

@tenant_admin_required
def all_employee_timesheets(request,  id, employee_id):
    # Retrieve the selected year and month from the GET parameters
    selected_year = request.GET.get('year')
    selected_month = request.GET.get('month')

    # Filter timesheets based on the selected year and month
    timesheets = Timesheets.objects.filter(employee_id=employee_id, company_id=request.user.company_id)

    if selected_year:
        timesheets = timesheets.filter(date__year=selected_year)

    if selected_month:
        timesheets = timesheets.filter(date__month=get_month_number(selected_month))

    employee = Employees.objects.get(id=id)

    # Define year and month ranges
    years = range(2025, 2027)
    # months = range(1, 13)
    months = ("Jan", "Feb","Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sept", "Oct", "Nov", "Dec")

    return render(request, "timesheets/all_employee_timesheets.html", {
        "timesheets": timesheets,
        "employee": employee,
        "years": years,
        "months": months,
        "year": selected_year,  # Pass the selected year back to the template
        "month": selected_month,  # Pass the selected month back to the template
    })




def is_clock_out(employee, date, site_id):
    clock_in = Timesheets.objects.filter(employee_id=employee, date=date, site_id=site_id)
    if clock_in.exists():
        return True
    else:
        return  False

    return  hours_worked

def get_hours_worked(clock_in, clock_out):
    # Combine date with time objects to create datetime objects
    date = datetime.today().date()  # Use any fixed date since we're only calculating time difference
    clock_in_dt = datetime.combine(date, clock_in)
    clock_out_dt = datetime.combine(date, clock_out)

    # If clock_out is earlier in the day than clock_in, assume clock_out is on the next day
    if clock_out_dt <= clock_in_dt:
        clock_out_dt += timedelta(days=1)

    # Calculate the difference in hours
    hours_worked = (clock_out_dt - clock_in_dt).total_seconds() / 3600

    # return math.floor(hours_worked*100)/100
    return  round(hours_worked, 2)





def get_year_month(date):
    return date.strftime("%Y-%m")

@tenant_admin_required
def upload(request):
    return render(request, 'timesheets/upload.html')

@tenant_admin_required
def timesheet_upload(request):
    time_sheets = Timesheets.objects.all()
    employees = Employees.objects.all()

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Read the Excel file using pandas

                df = pd.read_excel(file, header=0)
                df.columns = ['employee_id', 'clock_in', 'clock_out', 'date', 'site_id']

                df['date'] = pd.to_datetime(df['date'], errors='coerce')

                # Convert the 'clock_in' and 'clock_out' columns to time
                df['clock_in'] = pd.to_datetime(df['clock_in'], format='%H:%M:%S', errors='coerce').dt.time
                df['clock_out'] = pd.to_datetime(df['clock_out'], format='%H:%M:%S', errors='coerce').dt.time

                # Drop rows where date conversion failed
                df = df.dropna(subset=['date'])

                timesheets = []
                date_error =[]
                for index, row in df.iterrows():


                    work_day = Timesheets.objects.filter(employee_id=row['employee_id'],date=row['date'].date(), company_id=request.user.company_id)

                    if work_day.exists():
                        continue

                    hours_worked = get_hours_worked(row['clock_in'],row['clock_out'])
                    #
                    if hours_worked is None:
                        continue

                    employee = Employees.objects.filter(
                        employee_id=row['employee_id'],
                        company_id=request.user.company_id
                    ).first()

                    if not employee:
                        continue  # Skip if no matching employee in this company

                    timesheet = Timesheets(
                        company_id=request.user.company_id,
                        employee_id=row['employee_id'],
                        clock_in=row['clock_in'],
                        clock_out=row['clock_out'],
                        date=row['date'].date(),
                        site_id=employee.site_id if employee else None,  # safe lookup
                        hours_worked=hours_worked,
                        normal_hours=get_normal_hours(hours_worked, row['date'].date()),
                        overtime_normal_saturdays=get_overtime_one(request,row['clock_in'], row['clock_out'], row['date'].date()),
                        overtime_holiday_sundays=get_overtime_two(hours_worked, row['date'].date()),
                    )

                    timesheets.append(timesheet)



                # Bulk create TimeSheets objects
                Timesheets.objects.bulk_create(timesheets)
                messages.success(request, f"Timesheets uploaded successfully!")
            except Exception as e:
                print(e)
                messages.error(request, f"Error uploading timesheets: {e}")

            messages.success(request, f"{form}")
            return redirect('timesheets')

    else:
        form = UploadFileForm()

    return render(request, 'timesheets/upload_timesheet.html', {
        'form': form,
        'time_sheets': time_sheets
    })


@tenant_admin_required
def add_employee_timesheet(request, id):
    employee = Employees.objects.get(id=id)

    # Copy POST into mutable dict so we can transform it
    post_data = request.POST.copy()

    # Normalize date/time fields if they come in as Python objects
    for field in ["date", "clock_in", "clock_out"]:
        value = post_data.get(field)
        if isinstance(value, (dt_date, dt_time)):
            # Convert Python objects to strings Django form accepts
            if isinstance(value, dt_date):
                post_data[field] = value.strftime("%Y-%m-%d")
            elif isinstance(value, dt_time):
                post_data[field] = value.strftime("%H:%M")

    form = TimeSheetsForm(post_data or None, initial={'employee_id':employee.employee_id})

    if request.method == "POST":
        if form.is_valid():
            clock_in = form.cleaned_data['clock_in']
            clock_out = form.cleaned_data['clock_out']
            date = form.cleaned_data['date']
            hours_worked = get_hours_worked(clock_in, clock_out)

            timesheet = form.save(commit=False)  # don’t save yet
            timesheet.company = request.user.company

            timesheet.site = employee.site
            timesheet.hours_worked = hours_worked
            timesheet.normal_hours = get_normal_hours(hours_worked, date)
            timesheet.overtime_normal_saturdays = get_overtime_one(request,clock_in, clock_out, date)
            timesheet.overtime_holiday_sundays = get_overtime_two(hours_worked, date)
            timesheet.signed_off = 'Yes'
            timesheet.signed_off_by = request.user
            timesheet.save()  # ✅ now save with all fields

            messages.success(request, 'Form successfully added')
            return redirect('view_employee', id=id)
        else:
            print("FORM ERRORS:", form.errors)
            messages.error(request, 'Form is not valid')
            return redirect('view_employee', id=id)

    return render(request, "timesheets/add_employee_timesheet.html", {
        'form': form,
        'employee': employee
    })

@tenant_admin_required
def edit_employee_timesheet(request, id):
    timesheet = Timesheets.objects.get(id=id)
    # Copy POST into mutable dict so we can transform it
    post_data = request.POST.copy()

    # Normalize date/time fields if they come in as Python objects
    for field in ["date", "clock_in", "clock_out"]:
        value = post_data.get(field)
        if isinstance(value, (dt_date, dt_time)):
            # Convert Python objects to strings Django form accepts
            if isinstance(value, dt_date):
                post_data[field] = value.strftime("%Y-%m-%d")
            elif isinstance(value, dt_time):
                post_data[field] = value.strftime("%H:%M")
    if request.method == "POST":

        form = TimeSheetsForm(request.POST, request.FILES, instance=timesheet, user=request.user)  # Pass request.FILES here
        if form.is_valid():
            clock_in = form.cleaned_data['clock_in']
            clock_out = form.cleaned_data['clock_out']
            date = form.cleaned_data['date']
            hours_worked = get_hours_worked(clock_in, clock_out)

            timesheet = form.save(commit=False)  # don’t save yet


            timesheet.hours_worked = hours_worked
            timesheet.normal_hours = get_normal_hours(hours_worked, date)
            timesheet.overtime_normal_saturdays = get_overtime_one(request,clock_in, clock_out, date)
            timesheet.overtime_holiday_sundays = get_overtime_two(hours_worked, date)
            user = f'{request.user}'
            timesheet.signed_off_by = user
            timesheet.save()  # ✅ now save with all fields

            messages.success(request, f"Timesheet has been edited")
            return redirect('employees', )
    else:
        form = TimeSheetsForm(instance=timesheet, user=request.user)  # Correctly instantiate the form when GET request

    return render(request, "timesheets/edit_employee_timesheet.html", {
        'form': form,
        'timesheet': timesheet
    })


@login_required
def add_timesheet(request):
    current_time = datetime.datetime.now().time().replace(second=0, microsecond=0)

    form = TimeSheetsForm(initial={
                             'date': datetime.datetime.today().date(),
        # 'clock_in':current_time
    })
    if request.method == "POST":
        form = TimeSheetsForm(request.POST)
        if form.is_valid():

            employee = form.cleaned_data['employee']

            site = form.cleaned_data['site']

            if  is_clock_out(employee,  site):

                Timesheets.objects.filter(employee=employee,  site_id=site).update(clock_out=current_time)
                messages.success(request, f'{employee} clocked out at {current_time} from site {site}')
                return redirect('view_site_timesheets', site_id=site.site_id)
            else:
                form.save()
                messages.success(request, f'{employee} clocked in at {current_time} at {site}')
                return redirect('view_site_timesheets', site_id=site.site_id)

    # context = {'form': form}
    return render(request, "timesheets/add_timesheet.html", {
        'form': form
    })

def is_it_saturday_sunday(date):
    # Ensure the date is a datetime object
    if isinstance(date, str):
        date = datetime.strptime(date, "%Y-%m-%d").date()

    # Check day of the week
    if date.weekday() == 5:  # Saturday
        return 'Saturday'
    if date.weekday() == 6:
        return 'Sunday'
    if date.weekday() == 0:
        return 'Monday'
    if date.weekday() == 1:
        return 'Tuesday'
    if date.weekday() == 2:
        return 'Wednesday'
    if date.weekday() == 3:
        return 'Thursday'
    if date.weekday() == 4:
        return 'Friday'

    else:
        return None


def get_overtime_hours(hours_worked):
    if hours_worked is None:
        return 0
    if hours_worked <= 8:
        return 0
    if hours_worked > 8:
        overtime = hours_worked -8
        return round(overtime, 2)

def get_normal_hours(hours_worked,date):
    if hours_worked is None:
        return 0
    if is_it_saturday_sunday(date) == 'Saturday' or is_it_saturday_sunday(date) == 'Sunday':
        return 0
    if hours_worked > 9:
        return 8
    if hours_worked <= 9:
        return round(hours_worked, 2)


def is_public_holiday(date):
    return PublicHolidays.objects.filter(date=date).exists()


def get_overtime_one(request, clock_in, clock_out,date):
    company = request.user.company
    nightshift_hours = get_nightshift_hours(company, clock_in, clock_out)
    hours_worked = get_hours_worked(clock_in, clock_out)

    if is_it_saturday_sunday(date) == 'Saturday':
        return hours_worked
    elif nightshift_hours > 0:
        return round(nightshift_hours, 2)
    elif hours_worked > 9:  # typical SA normal working hours per day
        return round(hours_worked - 9, 2)
    else:
        return 0


def get_overtime_one_device(company, clock_in, clock_out,date):
    company = company
    nightshift_hours = get_nightshift_hours(company, clock_in, clock_out)
    hours_worked = get_hours_worked(clock_in, clock_out)

    if is_it_saturday_sunday(date) == 'Saturday':
        return hours_worked
    elif nightshift_hours > 0:
        return round(nightshift_hours, 2)
    elif hours_worked > 9:  # typical SA normal working hours per day
        return round(hours_worked - 9, 2)
    else:
        return 0

def get_nightshift_hours(company, clock_in, clock_out):


    if not company or not company.open or not company.close:
        return 0  # safety fallback

    open_time = company.open
    close_time = company.close

    if is_twentyfour_hours(open_time, close_time):
        return 0

    date = datetime.today().date()
    clock_in_dt = datetime.combine(date, clock_in)
    clock_out_dt = datetime.combine(date, clock_out)

    # Handle overnight shifts (e.g. 22:00 → 06:00 next day)
    if clock_out_dt <= clock_in_dt:
        clock_out_dt += timedelta(days=1)

    open_dt = datetime.combine(date, open_time)
    close_dt = datetime.combine(date, close_time)

    # Handle overnight business hours (e.g. open 22:00 → close 06:00)
    if close_dt <= open_dt:
        close_dt += timedelta(days=1)

    # Define the full 24-hour window for night hours
    night_start_1 = close_dt
    night_end_1 = open_dt + timedelta(days=1)

    # Calculate the overlap of [clock_in_dt, clock_out_dt] with [close_dt, next_day_open_dt]
    night_start = close_dt
    night_end = open_dt + timedelta(days=1)

    overlap_start = max(clock_in_dt, night_start)
    overlap_end = min(clock_out_dt, night_end)

    # Only count positive overlaps
    night_hours = 0
    if overlap_end > overlap_start:
        night_hours = (overlap_end - overlap_start).total_seconds() / 3600

    return round(night_hours, 2)


def is_twentyfour_hours(open_time, close_time):
    if open_time == close_time:
        return True
    return None


def get_overtime_two(hours_worked,date):
    if is_it_saturday_sunday(date) == 'Sunday' or is_public_holiday(date):
        return hours_worked
    else:
        return 0

pain = 'pain'




def get_owed_amount(request, wage, normal_hours, overtime_one, sunday_hours, holiday_hours):
    company = Company.objects.get(id=request.user.company_id)
    sunday_is_normal = company.sunday_is_normal_workday

    normal_fee = wage * normal_hours
    overtime_fee = (wage * overtime_one) * 1.5
    holiday_fee = (wage * holiday_hours) * 2

    if sunday_is_normal:
        sunday_fee = (wage * sunday_hours) * 1.5
    else:
        sunday_fee = (wage * sunday_hours) * 2

    total = normal_fee + overtime_fee + sunday_fee + holiday_fee
    return round(total, 2)



@login_required
def view_timesheet(request,  id):

    timesheet = Timesheets.objects.get(id=id)

    return render(request, "timesheets/view_timesheet.html", {
        'timesheet': timesheet,
        'overtime': get_overtime_hours(timesheet.hours_worked)


    })
@login_required
def sign_off_timesheet(request, id):

    timesheet = Timesheets.objects.get( id=id)
    sign_off = 'Yes'
    logged_in_user = request.user
    site = Timesheets.objects.get(id=id).site
    Timesheets.objects.filter(id=id,).update(signed_off =sign_off, signed_off_by=logged_in_user.username)


    messages.success(request, f'signed off')
    return redirect('view_site_employees', site_id=site)

def all_time_timesheets(request, employee_id):
    timesheet = Timesheets.objects.get(employee_id=employee_id)
    messages.success(request, f'signed off')
    return render(request, "timesheets/view_timesheet.html", {
        'timesheets': timesheets,

    })

@login_required
def delete_timesheet(request, id):
    timesheet = Timesheets.objects.get( id=id)
    timesheet.delete()
    messages.success(request, f'Removed')
    return redirect('timesheets')

def download_template(request):
    # Create a new workbook and active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set the headers for the Excel file
    headers = ['employee_id', 'clock_in', 'clock_out', 'date', 'site_id']
    worksheet.append(headers)

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="timesheet_template.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response


def payroll_data_excel(request, start_date, end_date):
    try:
        # Parse the start and end dates
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError as e:
        return render(request, "timesheets/error.html", {"message": f"Invalid date format: {e}"})

    # Filter timesheets by date range and signed-off status
    timesheets = Timesheets.objects.filter(
        date__range=[start_date, end_date],
        signed_off='Yes',
        company_id=request.user.company_id
    )

    # Aggregate employee data
    employee_hours = defaultdict(float)
    employee_normal_hours = defaultdict(float)
    employee_overtime_normal_saturdays = defaultdict(float)
    employee_sunday_hours = defaultdict(float)
    employee_public_holiday_hours = defaultdict(float)

    for timesheet in timesheets:
        employee_id = timesheet.employee_id
        employee_hours[employee_id] += timesheet.hours_worked
        employee_normal_hours[employee_id] += timesheet.normal_hours
        employee_overtime_normal_saturdays[employee_id] += timesheet.overtime_normal_saturdays

        # Separate Sunday vs Public Holiday
        if timesheet.date.weekday() == 6:  # Sunday
            employee_sunday_hours[employee_id] += timesheet.overtime_holiday_sundays
        elif is_public_holiday(timesheet.date):
            employee_public_holiday_hours[employee_id] += timesheet.overtime_holiday_sundays

    # Get employees for the company
    employees = Employees.objects.filter(company_id=request.user.company_id).values(
        'employee_id', 'first_name', 'last_name', 'position', 'site', 'wage'
    )

    # Create a new Excel workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payroll Data"

    # Set headers
    headers = [
        'Employee ID', 'First Name', 'Last Name', 'Position', 'Total Hours Worked',
        'Total Normal Hours', 'Overtime Saturdays', 'Sunday Hours', 'Public Holiday Hours',
        'Site Name', 'Amount Owed', 'Period'
    ]
    worksheet.append(headers)

    # Populate rows
    for employee in employees:
        employee_id = employee['employee_id']
        total_hours_worked = round(employee_hours.get(employee_id, 0), 2)
        if total_hours_worked > 0:
            normal_hours = employee_normal_hours.get(employee_id, 0)
            saturday_hours = employee_overtime_normal_saturdays.get(employee_id, 0)
            sunday_hours = employee_sunday_hours.get(employee_id, 0)
            public_holiday_hours = employee_public_holiday_hours.get(employee_id, 0)

            owed_amount = get_owed_amount(
                request,
                employee['wage'],
                normal_hours,
                saturday_hours,
                sunday_hours,
                public_holiday_hours
            )

            row = [
                employee_id,
                employee['first_name'],
                employee['last_name'],
                employee['position'],
                total_hours_worked,
                round(normal_hours, 2),
                saturday_hours,
                sunday_hours,
                public_holiday_hours,
                Sites.objects.get(site_id=employee['site']).site_name,
                owed_amount,
                f'{format_date(start_date)} - {format_date(end_date)}'
            ]
            worksheet.append(row)

    # Prepare Excel response
    filename = f'{format_date(start_date)} - {format_date(end_date)}_Payroll_Data.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response


@login_required
def bulk_sign_off_timesheets(request, employee_id, site_id):
    if request.method == 'POST':
        timesheet_ids = request.POST.getlist('timesheet_ids')
        logged_in_user = request.user

        if timesheet_ids:
            Timesheets.objects.filter(id__in=timesheet_ids).update(
                signed_off='Yes', signed_off_by=logged_in_user.username
            )
            messages.success(request, f"Selected timesheets signed off successfully!")
            return redirect('view_site_employees', site_id=site_id)
        else:
            messages.error(request, "No timesheets selected.")

    return redirect('view_employee_timesheets', employee_id=employee_id)


def bulk_sign_off_unknown_timesheets(request, employee_id, ):
    if request.method == 'POST':
        timesheet_ids = request.POST.getlist('timesheet_ids')
        logged_in_user = request.user

        if timesheet_ids:
            Timesheets.objects.filter(id__in=timesheet_ids).update(
                signed_off='Yes', signed_off_by=logged_in_user.username
            )
            messages.success(request, f"Selected timesheets signed off successfully!")
            return redirect('timesheets', )
        else:
            messages.error(request, "No timesheets selected.")

    return redirect('view_employee_timesheets', employee_id=employee_id)

@tenant_admin_required
def employees_upload(request):
    employees = Employees.objects.filter(company_id=request.user.company_id)
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Read the Excel file using pandas

                df = pd.read_excel(file, header=0)
                df.columns = ['Employee ID', 'First Name', 'Last Name', 'Wage']


                employees = []
                for index, row in df.iterrows():


                    employee = Employees.objects.filter(employee_id=row['Employee ID'], company_id=request.user.company_id)

                    if employee.exists():
                        continue

                    employee =  Employees(
                        employee_id=row['Employee ID'],
                        first_name=row['First Name'],
                        last_name=row['Last Name'],
                        wage=row['Wage'],
                        company_id=request.user.company_id,

                    )
                    employees.append(employee)



                # Bulk create TimeSheets objects
                Employees.objects.bulk_create(employees)
                messages.success(request, f"Employees uploaded successfully!")
            except Exception as e:
                print(e)
                messages.error(request, f"Error uploading timesheets: {e}")

            messages.success(request, f"{form}")
            return redirect('employees')

    else:
        form = UploadFileForm()

    return render(request, 'timesheets/upload_employees.html', {
        'form': form,
        'employees': employees
    })

def download_employees_template(request):
    # Create a new workbook and active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set the headers for the Excel file
    headers = ['Employee ID', 'First Name', 'Last Name', 'Wage']
    worksheet.append(headers)

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employees_template.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response
#Segoe UI', Tahoma, Geneva, Verdana, sans-serif